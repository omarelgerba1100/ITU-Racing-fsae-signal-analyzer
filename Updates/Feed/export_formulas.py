import os
import re
from openpyxl import load_workbook
import networkx as nx

# -------------------------------
# 0. Sanity check
# -------------------------------

print("Working directory:", os.getcwd())
print("Files in directory:", os.listdir("."))

EXCEL_FILE = "BEP.xlsm"

if EXCEL_FILE not in os.listdir("."):
    raise FileNotFoundError(f"{EXCEL_FILE} not found in current directory")

# -------------------------------
# 1. Load workbook
# -------------------------------

wb = load_workbook(EXCEL_FILE, data_only=False)

# -------------------------------
# 2. Export all formulas
# -------------------------------

with open("01_formulas.txt", "w", encoding="utf-8") as f:
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        f.write(f"\n=== Sheet: {sheet} ===\n")
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    f.write(f"{sheet}!{cell.coordinate} = {cell.value}\n")

print("Exported formulas.")

# -------------------------------
# 3. Export named ranges
# -------------------------------

with open("02_named_ranges.txt", "w", encoding="utf-8") as f:
    for name, defn in wb.defined_names.items():
        for sheet, ref in defn.destinations:
            f.write(f"{name} = {sheet}!{ref}\n")

print("Exported named ranges.")

# -------------------------------
# 4. Build dependency graph
# -------------------------------

cell_ref_global = re.compile(r"([A-Za-z0-9_]+)!([A-Z]+\d+)")
cell_ref_local = re.compile(r"\b([A-Z]+\d+)\b")

G = nx.DiGraph()

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                target = f"{sheet}!{cell.coordinate}"
                formula = cell.value

                # Cross-sheet references
                refs_global = cell_ref_global.findall(formula)
                for s, c in refs_global:
                    source = f"{s}!{c}"
                    G.add_edge(source, target)

                # Same-sheet references
                refs_local = cell_ref_local.findall(formula)
                for c in refs_local:
                    source = f"{sheet}!{c}"
                    if source != target:
                        G.add_edge(source, target)

with open("03_dependencies.txt", "w", encoding="utf-8") as f:
    for u, v in sorted(G.edges()):
        f.write(f"{u} -> {v}\n")

print("Exported dependency graph.")

# -------------------------------
# 5. Concatenate into markdown
# -------------------------------

files = [
    "01_formulas.txt",
    "02_named_ranges.txt",
    "03_dependencies.txt"
]

with open("excel_model.md", "w", encoding="utf-8") as out:
    for file in files:
        out.write(f"\n# {file}\n\n")
        with open(file, "r", encoding="utf-8") as f:
            out.write(f.read())

print("\nDONE.")
print("Final output: excel_model.md")
